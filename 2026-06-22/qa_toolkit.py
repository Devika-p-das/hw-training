import os
import sys
import argparse
import logging
from datetime import datetime
 
import numpy as np
import pandas as pd
 
# ── Optional: Excel output ────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
 
 
# ══════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════
 
logging.basicConfig(
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
)
log = logging.getLogger(__name__)
 
 
# ══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════
 
MISSING_PATTERNS = [
    "", " ", "  ",
    "NULL", "null", "Null",
    "NaN", "nan", "NAN",
    "N/A", "n/a", "NA", "na",
    "None", "none",
    "?", "-", "--",
    "unknown", "UNKNOWN", "Unknown",
    "missing", "MISSING",
]
 
 
# ══════════════════════════════════════════════════════════════════════
# FILE LOADING
# ══════════════════════════════════════════════════════════════════════
 
def load_file(path: str) -> pd.DataFrame:
    """Load any supported tabular file into a string DataFrame."""
 
    ext = os.path.splitext(path)[1].lower()
    log.info(f"Loading: {path}  ({ext})")
 
    loaders = {
        ".csv":     lambda: pd.read_csv(path, sep=None, engine="python", dtype=str),
        ".tsv":     lambda: pd.read_csv(path, sep="\t", dtype=str),
        ".xlsx":    lambda: pd.read_excel(path, dtype=str),
        ".xls":     lambda: pd.read_excel(path, dtype=str),
        ".json":    lambda: pd.read_json(path, dtype=str, lines=True),
        ".xml":     lambda: pd.read_xml(path, dtype=str),
        ".parquet": lambda: pd.read_parquet(path).astype(str),
    }
 
    if ext not in loaders:
        raise ValueError(f"Unsupported file type: {ext!r}. Supported: {list(loaders)}")
 
    df = loaders[ext]()
    log.info(f"Loaded {len(df):,} rows × {len(df.columns)} columns")
    return df
 
 
# ══════════════════════════════════════════════════════════════════════
# STANDARDIZE MISSING VALUES
# ══════════════════════════════════════════════════════════════════════
 
def standardize_missing(df: pd.DataFrame) -> pd.DataFrame:
    """Replace common missing-value strings with np.nan."""
    return df.replace(MISSING_PATTERNS, np.nan)
 
 
# ══════════════════════════════════════════════════════════════════════
# INDIVIDUAL CHECKS
# ══════════════════════════════════════════════════════════════════════
 
def check_basic_info(df: pd.DataFrame, file_path: str) -> dict:
    """File metadata — name, size, row/column counts, timestamp."""
 
    size_kb = round(os.path.getsize(file_path) / 1024, 2)
 
    return {
        "file_name":    os.path.basename(file_path),
        "file_size_kb": size_kb,
        "rows":         len(df),
        "columns":      len(df.columns),
        "run_at":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
 
 
def check_null_analysis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per-column null count, non-null count, null percentage.
    ⚠  flags columns with null% > 20.
    """
 
    rows = len(df)
    records = []
 
    for col in df.columns:
        null_count  = int(df[col].isna().sum())
        non_null    = int(df[col].notna().sum())
        null_pct    = round(null_count / rows * 100, 2) if rows else 0.0
 
        records.append({
            "column":   col,
            "null":     null_count,
            "non_null": non_null,
            "null_pct": null_pct,
        })
 
    return pd.DataFrame(records)
 
 
def check_uniqueness(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per-column distinct count, duplicate count, is_unique flag.
 
    BUG FIX vs v1: duplicate_count and is_unique now compare against
    non-null rows only, not total rows, so NaN rows don't inflate the counts.
    """
 
    records = []
 
    for col in df.columns:
        distinct    = int(df[col].nunique(dropna=True))
        non_null    = int(df[col].notna().sum())
        duplicates  = non_null - distinct        # ✅ fixed
        is_unique   = (distinct == non_null)     # ✅ fixed
 
        records.append({
            "column":     col,
            "distinct":   distinct,
            "duplicates": duplicates,
            "is_unique":  is_unique,
        })
 
    return pd.DataFrame(records)
 
 
def check_whitespace(df: pd.DataFrame) -> pd.DataFrame:
    """
    Detect leading / trailing whitespace in string columns.
    Returns only columns that actually have issues.
    """
 
    records = []
 
    for col in df.columns:
        series   = df[col].dropna().astype(str)
        leading  = int(series.str.startswith(" ").sum())
        trailing = int(series.str.endswith(" ").sum())
 
        if leading or trailing:
            records.append({
                "column":   col,
                "leading":  leading,
                "trailing": trailing,
            })
 
    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["column", "leading", "trailing"]
    )
 
 
def check_string_lengths(df: pd.DataFrame) -> pd.DataFrame:
    """Min / max / mean string length per column."""
 
    records = []
 
    for col in df.columns:
        lengths = df[col].dropna().astype(str).str.len()
        if lengths.empty:
            continue
 
        records.append({
            "column":   col,
            "min_len":  int(lengths.min()),
            "max_len":  int(lengths.max()),
            "mean_len": round(float(lengths.mean()), 1),
        })
 
    return pd.DataFrame(records)
 
 
def check_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    For every column that parses as numeric:
      min, max, mean, std, count of values ≤ 0,
      outlier count via IQR method (1.5 × IQR rule).
    """
 
    records = []
 
    for col in df.columns:
        num   = pd.to_numeric(df[col], errors="coerce")
        valid = num.dropna()
 
        if valid.empty:
            continue
 
        Q1, Q3  = valid.quantile([0.25, 0.75])
        IQR     = Q3 - Q1
        low     = Q1 - 1.5 * IQR
        high    = Q3 + 1.5 * IQR
        outliers = int(((valid < low) | (valid > high)).sum())
 
        records.append({
            "column":       col,
            "count":        int(valid.count()),
            "min":          round(float(valid.min()), 4),
            "max":          round(float(valid.max()), 4),
            "mean":         round(float(valid.mean()), 4),
            "std":          round(float(valid.std()), 4),
            "le_zero":      int((valid <= 0).sum()),
            "outliers_iqr": outliers,
        })
 
    return pd.DataFrame(records)
 
 
def check_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify columns that look like date/datetime values.
    Reports only columns where ≥ 50 % of non-null values parse as dates.
    """
 
    records = []
 
    for col in df.columns:
        parsed = pd.to_datetime(df[col], errors="coerce")
        valid  = int(parsed.notna().sum())
        total  = int(df[col].notna().sum())
 
        if total == 0:
            continue
 
        pct = round(valid / total * 100, 2)
 
        if pct >= 50:
            records.append({
                "column":        col,
                "parseable_pct": pct,
                "min_date":      str(parsed.min().date()) if valid else "N/A",
                "max_date":      str(parsed.max().date()) if valid else "N/A",
            })
 
    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["column", "parseable_pct", "min_date", "max_date"]
    )
 
 
def check_top_values(df: pd.DataFrame, top_n: int = 10) -> dict:
    """Top N value frequencies per column (including NaN)."""
 
    result = {}
 
    for col in df.columns:
        try:
            result[col] = df[col].value_counts(dropna=False).head(top_n)
        except Exception as e:
            log.warning(f"Top values failed for '{col}': {e}")
 
    return result
 
 
# ══════════════════════════════════════════════════════════════════════
# RUN ALL CHECKS
# ══════════════════════════════════════════════════════════════════════
 
def run_all_checks(df: pd.DataFrame, file_path: str, top_n: int = 10) -> dict:
    """Execute every check and return a results dictionary."""
 
    log.info("Running QA checks...")
 
    return {
        "basic":       check_basic_info(df, file_path),
        "null":        check_null_analysis(df),
        "uniqueness":  check_uniqueness(df),
        "whitespace":  check_whitespace(df),
        "str_lengths": check_string_lengths(df),
        "numeric":     check_numeric(df),
        "dates":       check_dates(df),
        "top_values":  check_top_values(df, top_n),
    }
 
 
# ══════════════════════════════════════════════════════════════════════
# TEXT REPORT
# ══════════════════════════════════════════════════════════════════════
 
SEP = "=" * 70
 
 
def _section(title: str) -> list:
    return ["\n", SEP, title, SEP]
 
 
def build_text_report(results: dict, df: pd.DataFrame) -> list:
    """Build a list of lines for the plain-text report."""
 
    lines = []
    b     = results["basic"]
 
    # Header
    lines += [SEP, "DATA QA REPORT", SEP]
    lines += [
        f"File      : {b['file_name']}",
        f"Size      : {b['file_size_kb']} KB",
        f"Rows      : {b['rows']:,}",
        f"Columns   : {b['columns']}",
        f"Run At    : {b['run_at']}",
    ]
 
    # ── Column names ──────────────────────────
    lines += _section("COLUMN NAMES")
    for i, col in enumerate(df.columns, 1):
        lines.append(f"  {i:>3}. {col}")
 
    # ── Empty columns ─────────────────────────
    lines += _section("COMPLETELY EMPTY COLUMNS")
    null_df = results["null"]
    empty   = null_df[null_df["null_pct"] == 100]["column"].tolist()
    lines  += ([f"  {c}" for c in empty] if empty else ["  None"])
 
    # ── Null analysis ─────────────────────────
    lines += _section("NULL ANALYSIS  (⚠ = null% > 20)")
    for _, row in results["null"].iterrows():
        flag = "  ⚠" if row["null_pct"] > 20 else ""
        lines.append(
            f"  {row['column']:<35} Null={row['null']:>6} | "
            f"NonNull={row['non_null']:>6} | Null%={row['null_pct']:>6}{flag}"
        )
 
    # ── Uniqueness ────────────────────────────
    lines += _section("UNIQUENESS ANALYSIS  (✓ = fully unique)")
    for _, row in results["uniqueness"].iterrows():
        flag = "  ✓" if row["is_unique"] else ""
        lines.append(
            f"  {row['column']:<35} Distinct={row['distinct']:>6} | "
            f"Duplicates={row['duplicates']:>6} | Unique={str(row['is_unique']):<5}{flag}"
        )
 
    # ── Whitespace ────────────────────────────
    lines += _section("WHITESPACE ISSUES")
    ws = results["whitespace"]
    if ws.empty:
        lines.append("  None detected")
    else:
        for _, row in ws.iterrows():
            lines.append(
                f"  {row['column']:<35} Leading={row['leading']:>4} | Trailing={row['trailing']:>4}"
            )
 
    # ── String lengths ────────────────────────
    lines += _section("STRING LENGTH STATS")
    for _, row in results["str_lengths"].iterrows():
        lines.append(
            f"  {row['column']:<35} Min={row['min_len']:>4} | "
            f"Max={row['max_len']:>4} | Mean={row['mean_len']:>6}"
        )
 
    # ── Numeric ───────────────────────────────
    lines += _section("NUMERIC STATS + OUTLIERS  (⚠ = IQR outliers found)")
    num = results["numeric"]
    if num.empty:
        lines.append("  No numeric columns detected")
    else:
        for _, row in num.iterrows():
            flag = f"  ⚠ {row['outliers_iqr']} outliers" if row["outliers_iqr"] > 0 else ""
            lines.append(
                f"  {row['column']:<35} Min={row['min']} | Max={row['max']} | "
                f"Mean={row['mean']} | Std={row['std']} | <=0={row['le_zero']}{flag}"
            )
 
    # ── Dates ─────────────────────────────────
    lines += _section("DATE COLUMN DETECTION")
    dates = results["dates"]
    if dates.empty:
        lines.append("  No date-like columns detected")
    else:
        for _, row in dates.iterrows():
            lines.append(
                f"  {row['column']:<35} Parseable={row['parseable_pct']}% | "
                f"Range: {row['min_date']} → {row['max_date']}"
            )
 
    # ── Top values ────────────────────────────
    lines += _section("TOP VALUES  (per column)")
    for col, counts in results["top_values"].items():
        lines.append(f"\n  COLUMN: {col}")
        for val, cnt in counts.items():
            lines.append(f"    {str(val):<45} : {cnt}")
 
    return lines
 
 
def save_text_report(lines: list, out_path: str):
    with open(out_path, "w", encoding="utf-8") as f:
        for line in lines:
            f.write(str(line) + "\n")
    log.info(f"TXT  report saved → {out_path}")
 
 
# ══════════════════════════════════════════════════════════════════════
# EXCEL REPORT  (color-coded)
# ══════════════════════════════════════════════════════════════════════
 
def save_excel_report(results: dict, out_path: str):
 
    if not EXCEL_OK:
        log.warning("openpyxl not installed — skipping Excel output.  Run: pip install openpyxl")
        return
 
    # ── Color palette ─────────────────────────
    C_HEADER  = PatternFill("solid", fgColor="2C3E50")   # dark slate
    C_RED     = PatternFill("solid", fgColor="FF6B6B")   # > 50% null / problem
    C_YELLOW  = PatternFill("solid", fgColor="FFE66D")   # warning
    C_GREEN   = PatternFill("solid", fgColor="A8E6CF")   # good / unique
    C_BLUE    = PatternFill("solid", fgColor="C8E6FF")   # date columns
 
    F_WHITE   = Font(bold=True, color="FFFFFF")
    F_BOLD    = Font(bold=True)
 
    # ── Helper: write a DataFrame to a sheet ──
    def write_df(ws, dataframe: pd.DataFrame):
        headers = list(dataframe.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = C_HEADER
            cell.font      = F_WHITE
            cell.alignment = Alignment(horizontal="center")
        for row_data in dataframe.itertuples(index=False):
            ws.append(list(row_data))
        # auto-column width
        for col_idx, _ in enumerate(headers, 1):
            letter = get_column_letter(col_idx)
            max_w  = max(
                len(str(cell.value or ""))
                for cell in ws[letter]
            )
            ws.column_dimensions[letter].width = min(max_w + 4, 55)
 
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet
 
    # ── Sheet 1 : Summary ─────────────────────
    ws = wb.create_sheet("Summary")
    ws.append(["Attribute", "Value"])
    for cell in ws[1]:
        cell.fill = C_HEADER
        cell.font = F_WHITE
    for k, v in results["basic"].items():
        ws.append([k, str(v)])
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            cell.font = F_BOLD
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
 
    # ── Sheet 2 : Null Analysis ───────────────
    ws = wb.create_sheet("Null Analysis")
    write_df(ws, results["null"])
    for row in ws.iter_rows(min_row=2):
        try:
            pct = float(row[3].value)      # null_pct
        except (TypeError, ValueError):
            continue
        if pct > 50:
            color = C_RED
        elif pct > 20:
            color = C_YELLOW
        elif pct == 0:
            color = C_GREEN
        else:
            color = None
        if color:
            for cell in row:
                cell.fill = color
 
    # ── Sheet 3 : Uniqueness ──────────────────
    ws = wb.create_sheet("Uniqueness")
    write_df(ws, results["uniqueness"])
    for row in ws.iter_rows(min_row=2):
        is_uniq = row[3].value           # is_unique
        if is_uniq is True or str(is_uniq).lower() == "true":
            for cell in row:
                cell.fill = C_GREEN
 
    # ── Sheet 4 : Whitespace ──────────────────
    ws = wb.create_sheet("Whitespace")
    if results["whitespace"].empty:
        ws.append(["✓ No whitespace issues found"])
        ws["A1"].fill = C_GREEN
    else:
        write_df(ws, results["whitespace"])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = C_YELLOW
 
    # ── Sheet 5 : String Lengths ──────────────
    ws = wb.create_sheet("String Lengths")
    write_df(ws, results["str_lengths"])
 
    # ── Sheet 6 : Numeric Stats ───────────────
    ws = wb.create_sheet("Numeric Stats")
    num = results["numeric"]
    if num.empty:
        ws.append(["No numeric columns detected"])
    else:
        write_df(ws, num)
        for row in ws.iter_rows(min_row=2):
            try:
                outliers = int(row[7].value)   # outliers_iqr
            except (TypeError, ValueError):
                continue
            if outliers > 0:
                for cell in row:
                    cell.fill = C_YELLOW
 
    # ── Sheet 7 : Date Columns ────────────────
    ws = wb.create_sheet("Date Columns")
    if results["dates"].empty:
        ws.append(["No date-like columns detected"])
    else:
        write_df(ws, results["dates"])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = C_BLUE
 
    # ── Sheet 8 : Top Values ──────────────────
    ws = wb.create_sheet("Top Values")
    ws.append(["Column", "Value", "Count"])
    for cell in ws[1]:
        cell.fill = C_HEADER
        cell.font = F_WHITE
    for col, counts in results["top_values"].items():
        for val, cnt in counts.items():
            ws.append([col, str(val), int(cnt)])
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
 
    wb.save(out_path)
    log.info(f"Excel report saved → {out_path}")
 
 
# ══════════════════════════════════════════════════════════════════════
# PUBLIC API  (import-friendly for notebooks / other scripts)
# ══════════════════════════════════════════════════════════════════════
 
def profile(file_path: str, top_n: int = 10, output: str = "both") -> dict:
    """
    One-call API — load, clean, check, save.
 
    Parameters
    ----------
    file_path : str   path to the input file
    top_n     : int   top N values per column  (default 10)
    output    : str   'txt' | 'excel' | 'both' (default 'both')
 
    Returns
    -------
    dict  raw results (use for further analysis in notebooks)
 
    Example
    -------
    from qa_toolkit import profile
    results = profile("sales_data.csv", output="both")
    print(results["null"])
    """
 
    df      = load_file(file_path)
    df      = standardize_missing(df)
    results = run_all_checks(df, file_path, top_n)
    base    = os.path.splitext(file_path)[0]
 
    if output in ("txt", "both"):
        lines = build_text_report(results, df)
        save_text_report(lines, f"{base}_qa_report.txt")
 
    if output in ("excel", "both"):
        save_excel_report(results, f"{base}_qa_report.xlsx")
 
    return results
 
 
# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════
 
def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Data QA Toolkit — profile any tabular data file",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
  python qa_toolkit.py sales.csv
  python qa_toolkit.py data.xlsx  --output excel
  python qa_toolkit.py dump.csv   --output both --top 20
""",
    )
    p.add_argument("file",
                   help="Path to the input file")
    p.add_argument("--output", choices=["txt", "excel", "both"], default="txt",
                   help="Output format (default: txt)")
    p.add_argument("--top",    type=int, default=10,
                   help="Top N values per column (default: 10)")
    return p.parse_args()
 
 
def main():
    args = _parse_args()
 
    if not os.path.exists(args.file):
        log.error(f"File not found: {args.file}")
        sys.exit(1)
 
    profile(args.file, top_n=args.top, output=args.output)
    log.info("Done.")
 
 
if __name__ == "__main__":
    main()
 
